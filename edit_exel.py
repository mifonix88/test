import sys
import subprocess
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

path_to_file = sys.argv[1]#получаем путь к открываемому файлу
path_to_exel = 'D:/Office16/EXCEL.EXE'

all_katalogy = [r'C:\Users\user\YandexDisk\В работу\В работу\OZON\Штрихкоды_ИП',
                r'C:\Users\user\YandexDisk\В работу\В работу\OZON\Штрихкоды_ООО']

def all_files(adresa):
    '''
    вернёт словарь {имя файла:путь к нему} из all_katalogy()[arg]
    '''
    res = {}
    for adres in adresa:
        for i in os.walk(adres): 
            for j in i[2]:
                res[j] = f'{adres}\{j}'

    return res



def name_search(list_names, arg):
    '''
    вернёт список имён из list_names удовлетворяющих поиску(ищит arg)
    '''
    temp_res = []
    res = []
    arg = [i.lower().strip() for i in arg.split()] 



    for j in list_names:
        if len(arg)==1 and arg[0].isdigit(): #если аргумент один и состоит только из цифр
            try:
                if int(arg[0]) == int(j[:len(arg[0])]) or int(arg[0]) == int(j[:3]): #проверям не начинается ли с него строка
                    res.append(j)
            except(ValueError):pass

            i = j.strip().lower()
            if i.find(arg[0]) > 3:
                temp_res.append(j)
        else:
            temp = 0
            i = j.strip().lower()

            for k in arg:
                if i.find(k) > -1:
                    temp +=1
            if temp == len(arg):
                res.append(j)

    return res if res else temp_res








try:
    wb = load_workbook(path_to_file)

    #первый лист
    w_list = wb[wb.sheetnames[0]]


    #находим намера столбцов с артиклом и наименованием
    names = list(w_list.values)[0]
    art = names.index('Артикул продавца')
    name = names.index('Наименование')

    def intersection(arg):
        '''
        Принимает итерируемый объект.
        Возвращает словать {наименование(str):количество повторений(int)} 
        '''
        res = {}
        for i in arg:
            if res.get(i):
                res[i]+=1
            else:
                res[i] = 1
        return res

    #получаем искомые данные
    temp_val = [(i[art],i[name]) for i in  w_list.values]
    value = intersection(temp_val[1:])


    #удаляем страницу
    #wb.remove(wb[wb.sheetnames[0]])

    #создаём новую заполняем
    ws = wb.create_sheet("Общий", 0)
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['A'].width = 20


    def _top(ls, ws):
        for i in ls:
            arg = ws.cell(row=1, column=1+ls.index(i), value = i)
            arg.font = Font(size= 12, bold=True)
            arg.alignment = Alignment(horizontal='center')


    ls = ('Арт.', 'Наименование', 'Всего:')
    _top(ls, ws)


    name_files = all_files(all_katalogy)


    for index, _art in enumerate(value):

        index = index + 2
                
        art = _art[0]
        try:
            name = name_search(name_files, art)[0].split('.')
            name = ''.join(name[:-1])
            name = name.split('_')
            name = '_'.join(name[1:])
        except:
            name = _art[1]
        
        
        
        
        kol = value[_art]
        
        arg1 = ws.cell(row=index, column=1, value = art)
        arg1.alignment = Alignment(horizontal='center')
                
        arg2 = ws.cell(row=index, column=2, value = name)
        arg2.alignment = Alignment(horizontal='left')

        arg3 = ws.cell(row=index, column=3, value = kol)
        arg3.alignment = Alignment(horizontal='center')
        arg3.font = Font(bold=True)


    #сохраняем
    wb.save(path_to_file)


    

    
except:pass
finally: subprocess.Popen([path_to_exel, path_to_file])

